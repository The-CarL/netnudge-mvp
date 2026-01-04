from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .. import OutreachRecord

HEADERS = [
    "Name",
    "Company",
    "Phone",
    "Email",
    "LinkedIn URL",
    "Channel",
    "Match Confidence",
    "Message",
    "Sent",
]

COLUMN_WIDTHS = {
    "Name": 25,
    "Company": 25,
    "Phone": 18,
    "Email": 30,
    "LinkedIn URL": 40,
    "Channel": 12,
    "Match Confidence": 18,
    "Message": 60,
    "Sent": 8,
}


def write_xlsx(records: list[OutreachRecord], output_path: str | Path) -> Path:
    """
    Write outreach records to an Excel spreadsheet.

    Args:
        records: List of OutreachRecord objects to write
        output_path: Path to save the Excel file

    Returns:
        Path to the created file
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Outreach"

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # Write headers
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row_idx, record in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1, value=record.name)
        ws.cell(row=row_idx, column=2, value=record.company or "")
        ws.cell(row=row_idx, column=3, value=record.phone or "")
        ws.cell(row=row_idx, column=4, value=record.email or "")
        ws.cell(row=row_idx, column=5, value=record.linkedin_url or "")
        ws.cell(row=row_idx, column=6, value=record.channel)
        ws.cell(row=row_idx, column=7, value=record.match_confidence)

        # Message cell with wrap
        msg_cell = ws.cell(row=row_idx, column=8, value=record.message)
        msg_cell.alignment = wrap_alignment

        # Sent checkbox (FALSE = unchecked)
        ws.cell(row=row_idx, column=9, value=record.sent)

    # Set column widths
    for col_idx, header in enumerate(HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(header, 15)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add autofilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    wb.save(output_path)
    return output_path
